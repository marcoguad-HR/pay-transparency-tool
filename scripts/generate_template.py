#!/usr/bin/env python3
"""
Generate Template — Crea il file Excel template per l'analisi gender pay gap.

Genera un file .xlsx con 3 fogli:
1. Template: header corretti + data validation + 3 righe esempio
2. Esempio: 20 righe di dati fittizi realistici
3. Istruzioni: guida step-by-step in italiano

Uso:
    python scripts/generate_template.py
    python scripts/generate_template.py --output /path/to/output.xlsx
"""

import argparse
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
    sys.exit(1)

OUTPUT_PATH = project_root / "static" / "template_retribuzioni.xlsx"

# Colori
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")  # Light indigo
OPTIONAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Light gray
EXAMPLE_FONT = Font(name="Calibri", color="6B7280", size=10, italic=True)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COLUMNS = [
    # (header, width, required, description)
    ("employee_id", 14, False, "Identificativo dipendente (opzionale)"),
    ("gender", 10, True, "Genere: M o F"),
    ("department", 18, False, "Dipartimento (consigliato per analisi per categoria)"),
    ("level", 14, False, "Livello aziendale (consigliato per analisi per categoria)"),
    ("base_salary", 16, True, "Stipendio lordo annuo in euro (senza simbolo)"),
    ("bonus", 14, False, "Premio variabile annuo in euro (opzionale)"),
]

EXAMPLE_DATA = [
    ("EMP001", "F", "Engineering", "Senior", 48000, 5000),
    ("EMP002", "M", "Engineering", "Senior", 52000, 6200),
    ("EMP003", "F", "Engineering", "Junior", 32000, 2000),
    ("EMP004", "M", "Engineering", "Junior", 33500, 2500),
    ("EMP005", "F", "Marketing", "Mid", 38000, 3000),
    ("EMP006", "M", "Marketing", "Mid", 41000, 3800),
    ("EMP007", "F", "Marketing", "Senior", 50000, 5500),
    ("EMP008", "M", "Marketing", "Senior", 53000, 6000),
    ("EMP009", "F", "Sales", "Junior", 30000, 4000),
    ("EMP010", "M", "Sales", "Junior", 31000, 4500),
    ("EMP011", "F", "Sales", "Senior", 45000, 8000),
    ("EMP012", "M", "Sales", "Senior", 49000, 9500),
    ("EMP013", "F", "Finance", "Mid", 42000, 3500),
    ("EMP014", "M", "Finance", "Mid", 44000, 4000),
    ("EMP015", "F", "HR", "Junior", 29000, 1500),
    ("EMP016", "M", "HR", "Junior", 29500, 1800),
    ("EMP017", "F", "HR", "Senior", 43000, 4000),
    ("EMP018", "M", "HR", "Senior", 44000, 4200),
    ("EMP019", "F", "Operations", "Mid", 36000, 2500),
    ("EMP020", "M", "Operations", "Mid", 37500, 2800),
]

TEMPLATE_EXAMPLES = [
    ("", "F", "Engineering", "Senior", 48000, 5000),
    ("", "M", "Sales", "Junior", 31000, 4500),
    ("", "F", "HR", "Mid", 38000, ""),
]

INSTRUCTIONS = [
    ("Come preparare i dati per l'analisi Gender Pay Gap", "title"),
    ("", "spacer"),
    ("1. COLONNE OBBLIGATORIE", "section"),
    ("gender — Il genere del dipendente. Valori accettati: M (uomo) o F (donna). Non case-sensitive.", "text"),
    ("base_salary — Lo stipendio lordo annuo in euro. Inserire solo il numero, senza simbolo € o separatori delle migliaia.", "text"),
    ("", "spacer"),
    ("2. COLONNE CONSIGLIATE", "section"),
    ("department — Il dipartimento di appartenenza (es. Engineering, Marketing, HR). Permette l'analisi del gap per categoria, obbligatoria per la compliance EU.", "text"),
    ("level — Il livello aziendale (es. Junior, Mid, Senior, Lead). Insieme a department, permette il breakdown dettagliato richiesto dalla Direttiva EU 2023/970.", "text"),
    ("", "spacer"),
    ("3. COLONNE OPZIONALI", "section"),
    ("employee_id — Un identificativo univoco per ogni dipendente. Utile per il tuo tracking interno, non usato nell'analisi.", "text"),
    ("bonus — Il premio variabile annuo lordo in euro. Se presente, il tool calcola anche il gap sulla retribuzione variabile.", "text"),
    ("", "spacer"),
    ("4. COME ESPORTARE DA EXCEL", "section"),
    ("1. Apri il foglio 'Template' e compila i dati dei tuoi dipendenti", "text"),
    ("2. File > Salva con nome > scegli formato 'CSV UTF-8 (delimitato da virgola)'", "text"),
    ("3. In alternativa, salva direttamente come .xlsx e caricalo sul tool", "text"),
    ("", "spacer"),
    ("5. NOTE IMPORTANTI", "section"),
    ("- Servono almeno 2 dipendenti per genere per calcolare il gap", "text"),
    ("- Si consigliano almeno 50 dipendenti per risultati statisticamente significativi", "text"),
    ("- Per l'analisi per categoria servono almeno 2 persone per genere in ogni gruppo department+level", "text"),
    ("- I gruppi con meno di 2 persone per genere vengono esclusi dall'analisi per categoria", "text"),
    ("- La soglia EU per il gap retributivo e' del 5%: se superata, e' richiesta una valutazione congiunta (Art. 9, Direttiva 2023/970)", "text"),
    ("", "spacer"),
    ("6. PRIVACY", "section"),
    ("- Sul portale web: i dati vengono elaborati in memoria e cancellati subito dopo. Non vengono salvati.", "text"),
    ("- Con il tool offline: i dati non lasciano MAI il tuo computer. L'analisi avviene interamente nel browser.", "text"),
]


def _style_header(ws, row: int, ncols: int):
    """Applica stile all'header."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def create_template(output_path: Path):
    wb = Workbook()

    # =========================================================================
    # FOGLIO 1: Template
    # =========================================================================
    ws_template = wb.active
    ws_template.title = "Template"
    ws_template.sheet_properties.tabColor = "4F46E5"

    # Riga descrizione
    ws_template.merge_cells("A1:F1")
    desc_cell = ws_template["A1"]
    desc_cell.value = "Compila i dati dei tuoi dipendenti. Le colonne viola sono obbligatorie."
    desc_cell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    desc_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_template.row_dimensions[1].height = 25

    # Header (riga 2)
    for col_idx, (name, width, required, desc) in enumerate(COLUMNS, 1):
        cell = ws_template.cell(row=2, column=col_idx, value=name)
        ws_template.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws_template, 2, len(COLUMNS))

    # Sub-header con indicazione obbligatorio/opzionale (riga 3)
    for col_idx, (name, width, required, desc) in enumerate(COLUMNS, 1):
        cell = ws_template.cell(row=3, column=col_idx)
        cell.value = "OBBLIGATORIO" if required else "opzionale"
        cell.font = Font(name="Calibri", size=9, bold=required, color="4F46E5" if required else "9CA3AF")
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Righe esempio (riga 4-6)
    for row_idx, row_data in enumerate(TEMPLATE_EXAMPLES, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_template.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    # Data validation su gender
    dv_gender = DataValidation(
        type="list",
        formula1='"M,F"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valore non valido",
        error="Inserire M (uomo) o F (donna)",
    )
    dv_gender.add("B4:B1000")
    ws_template.add_data_validation(dv_gender)

    # Data validation su base_salary (numero > 0)
    dv_salary = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valore non valido",
        error="Inserire un numero intero positivo (stipendio lordo annuo in euro)",
    )
    dv_salary.add("E4:E1000")
    ws_template.add_data_validation(dv_salary)

    # Freeze panes
    ws_template.freeze_panes = "A4"

    # =========================================================================
    # FOGLIO 2: Esempio
    # =========================================================================
    ws_example = wb.create_sheet("Esempio (20 dipendenti)")
    ws_example.sheet_properties.tabColor = "10B981"

    # Header
    for col_idx, (name, width, required, desc) in enumerate(COLUMNS, 1):
        ws_example.cell(row=1, column=col_idx, value=name)
        ws_example.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws_example, 1, len(COLUMNS))

    # Dati
    for row_idx, row_data in enumerate(EXAMPLE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_example.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    ws_example.freeze_panes = "A2"

    # =========================================================================
    # FOGLIO 3: Istruzioni
    # =========================================================================
    ws_instructions = wb.create_sheet("Istruzioni")
    ws_instructions.sheet_properties.tabColor = "F59E0B"
    ws_instructions.column_dimensions["A"].width = 100

    row = 1
    for text, style in INSTRUCTIONS:
        cell = ws_instructions.cell(row=row, column=1, value=text)
        if style == "title":
            cell.font = Font(name="Calibri", size=14, bold=True, color="1F2937")
            ws_instructions.row_dimensions[row].height = 30
        elif style == "section":
            cell.font = Font(name="Calibri", size=12, bold=True, color="4F46E5")
            ws_instructions.row_dimensions[row].height = 24
        elif style == "text":
            cell.font = Font(name="Calibri", size=11, color="374151")
            cell.alignment = Alignment(wrap_text=True)
        elif style == "spacer":
            ws_instructions.row_dimensions[row].height = 10
        row += 1

    # Salva
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Template salvato in: {output_path}")
    print(f"Dimensione: {output_path.stat().st_size / 1024:.1f} KB")


def main():
    parser = argparse.ArgumentParser(description="Genera template Excel per analisi gender pay gap")
    parser.add_argument("--output", type=str, default=str(OUTPUT_PATH), help=f"Path di output (default: {OUTPUT_PATH})")
    args = parser.parse_args()

    create_template(Path(args.output))


if __name__ == "__main__":
    main()
